"""RFP requirement extraction and compliance matrix.

    python rfp_compliance.py extract rfp.txt --out requirements.csv
    python rfp_compliance.py matrix requirements.csv [--xlsx Compliance_Matrix.xlsx]

`extract` is a first pass, never the answer: it finds obligation sentences (shall / must /
required to / should / may) and keeps the nearest section number. It will miss requirements in
tables, annexes and evaluation criteria, and it will over-capture boilerplate. Review every row.

For Word or PDF source documents, extract the text first with
`07-contract-review/scripts/contract_extract.py`, then run `extract` over the result.

Standard library only; `--xlsx` additionally needs openpyxl.
"""

import argparse
import csv
import os
import re
import sys

# Requirement text is quoted verbatim and often carries non-ASCII punctuation; Windows
# consoles and redirected output default to cp1252 and would raise UnicodeEncodeError.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

MANDATORY_WORDS = ["shall", "must", "is required to", "are required to", "will be required"]
DESIRABLE_WORDS = ["should", "may", "is preferred", "desirable"]
SECTION_RE = re.compile(r"^\s*((?:\d+\.)+\d*|\d+|[A-Z]\.\d+)\s+\S")
STATUSES = ["compliant", "partial", "alternative", "non-compliant", ""]

COLUMNS = ["req_id", "section", "requirement", "type", "weight", "owner", "compliance",
           "response_location", "evidence", "notes"]


# ---------------------------------------------------------------- extract

def split_sentences(text):
    parts = re.split(r"(?<=[.;])\s+(?=[A-Z(])", text)
    return [p.strip() for p in parts if p.strip()]


def classify(sentence):
    lowered = sentence.lower()
    for word in MANDATORY_WORDS:
        if word in lowered:
            return "mandatory"
    for word in DESIRABLE_WORDS:
        if re.search(r"\b" + re.escape(word) + r"\b", lowered):
            return "desirable"
    return None


def run_extract(args):
    with open(args.source, encoding="utf-8", errors="replace") as fh:
        raw = fh.read()

    rows = []
    section = ""
    seen = set()
    for line in raw.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        match = SECTION_RE.match(line)
        if match:
            section = match.group(1).rstrip(".")
        for sentence in split_sentences(stripped):
            kind = classify(sentence)
            if not kind:
                continue
            if len(sentence) < args.min_length or len(sentence) > args.max_length:
                continue
            key = sentence.lower()
            if key in seen:
                continue
            seen.add(key)
            # Drop a leading section number already captured in its own column.
            if section and sentence.startswith(section):
                sentence = sentence[len(section):].lstrip(". \t")
            rows.append({
                "req_id": "R-{:03d}".format(len(rows) + 1),
                "section": section,
                "requirement": sentence,
                "type": kind,
                "weight": "",
                "owner": "",
                "compliance": "",
                "response_location": "",
                "evidence": "",
                "notes": "",
            })

    with open(args.out, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)

    mandatory = sum(1 for r in rows if r["type"] == "mandatory")
    print("extracted {} candidate requirements ({} mandatory, {} desirable) -> {}".format(
        len(rows), mandatory, len(rows) - mandatory, args.out))
    print("\nNEXT: review every row by hand. The parser cannot see requirements in tables, "
          "annexes or evaluation criteria, and it over-captures boilerplate.")
    print("Then set owner, compliance and response_location on each row and run:")
    print("    python rfp_compliance.py matrix {}".format(args.out))
    return 0


# ---------------------------------------------------------------- matrix

def read_requirements(path):
    with open(path, newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))
    for row in rows:
        row["type"] = (row.get("type") or "").strip().lower()
        row["compliance"] = (row.get("compliance") or "").strip().lower()
    return rows


def run_matrix(args):
    rows = read_requirements(args.requirements)
    if not rows:
        sys.exit("no requirements found in " + args.requirements)

    mandatory = [r for r in rows if r["type"] == "mandatory"]
    desirable = [r for r in rows if r["type"] != "mandatory"]

    def count(subset, status):
        return sum(1 for r in subset if r["compliance"] == status)

    open_mandatory = [r for r in mandatory if r["compliance"] not in ("compliant", "alternative")]
    unassigned = [r for r in rows if not (r.get("owner") or "").strip()]
    unplaced = [r for r in rows if r["compliance"] in ("compliant", "partial", "alternative")
                and not (r.get("response_location") or "").strip()]
    no_evidence = [r for r in rows if r["compliance"] == "compliant"
                   and not (r.get("evidence") or "").strip()]
    bad_status = [r for r in rows if r["compliance"] not in STATUSES]

    def weight(subset):
        total = 0.0
        for r in subset:
            try:
                total += float((r.get("weight") or "0").strip() or 0)
            except ValueError:
                pass
        return total

    weight_at_risk = weight([r for r in rows if r["compliance"] in ("", "partial", "non-compliant")])

    lines = ["# Compliance matrix", "",
             "| Req | Section | Type | Wt | Compliance | Owner | Response location | Evidence |",
             "|---|---|---|---|---|---|---|---|"]
    for r in rows:
        lines.append("| {} | {} | {} | {} | {} | {} | {} | {} |".format(
            r.get("req_id", ""), r.get("section", ""), r.get("type", ""), r.get("weight", ""),
            r["compliance"] or "**OPEN**", r.get("owner", "") or "**UNASSIGNED**",
            r.get("response_location", ""), (r.get("evidence", "") or "")[:70]))

    lines += ["", "## Coverage", "",
              "| | Total | Compliant | Partial | Alternative | Non-compliant | Open |",
              "|---|---|---|---|---|---|---|"]
    for label, subset in [("Mandatory", mandatory), ("Desirable", desirable), ("All", rows)]:
        lines.append("| {} | {} | {} | {} | {} | {} | {} |".format(
            label, len(subset), count(subset, "compliant"), count(subset, "partial"),
            count(subset, "alternative"), count(subset, "non-compliant"), count(subset, "")))

    lines += ["", "Evaluation weight currently unproven (open, partial or non-compliant): "
                  "**{:g}**".format(weight_at_risk), ""]

    lines += ["## Blocking checks", ""]
    if open_mandatory:
        lines.append("**{} mandatory requirement(s) not yet compliant - nothing else proceeds "
                     "until this list is empty:**".format(len(open_mandatory)))
        for r in open_mandatory:
            lines.append("- `{}` ({}) [{}] {}".format(
                r.get("req_id"), r.get("section"), r["compliance"] or "OPEN",
                (r.get("requirement") or "")[:120]))
    else:
        lines.append("- All mandatory requirements are compliant or have a recorded alternative.")

    for label, subset, note in [
            ("Unassigned (no owner)", unassigned, "assign an owner to every requirement"),
            ("No response location", unplaced, "state where in the response each is answered"),
            ("Compliant but no evidence", no_evidence,
             "a compliant claim without evidence scores mid-band at best"),
            ("Invalid compliance status", bad_status,
             "use compliant / partial / alternative / non-compliant")]:
        if subset:
            lines.append("")
            lines.append("**{}** ({}) — {}:".format(label, len(subset), note))
            for r in subset[:20]:
                lines.append("- `{}` {}".format(r.get("req_id"), (r.get("requirement") or "")[:100]))
            if len(subset) > 20:
                lines.append("- ... and {} more".format(len(subset) - 20))

    if any(r["compliance"] == "alternative" for r in rows):
        lines += ["", "> `alternative` positions must be agreed internally and flagged clearly in "
                      "the response. An unflagged deviation reads as non-compliance to an evaluator."]
    if any(r["compliance"] == "non-compliant" for r in rows):
        lines += ["", "> Non-compliant items on commercial or liability terms: route to the "
                      "`contract-review` skill (07) before submission."]

    markdown = "\n".join(lines)
    print(markdown)

    if args.out:
        with open(args.out, "w", encoding="utf-8") as fh:
            fh.write(markdown)
        print("\nwrote " + args.out)

    if args.xlsx:
        write_xlsx(args.xlsx, rows, open_mandatory, weight_at_risk)
        print("wrote " + args.xlsx)

    return 1 if open_mandatory else 0


def write_xlsx(path, rows, open_mandatory, weight_at_risk):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        sys.exit("--xlsx needs openpyxl:  pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Matrix"
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    fills = {"compliant": PatternFill("solid", fgColor="C6EFCE"),
             "partial": PatternFill("solid", fgColor="FFEB9C"),
             "alternative": PatternFill("solid", fgColor="DDEBF7"),
             "non-compliant": PatternFill("solid", fgColor="FFC7CE"),
             "": PatternFill("solid", fgColor="F2F2F2")}

    ws.append([c.replace("_", " ").title() for c in COLUMNS])
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for width, col in zip([9, 10, 76, 12, 7, 16, 15, 22, 46, 34], range(1, len(COLUMNS) + 1)):
        ws.column_dimensions[get_column_letter(col)].width = width

    for r in rows:
        ws.append([r.get(c, "") for c in COLUMNS])
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=7).fill = fills.get(r["compliance"], fills[""])
        ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_idx, column=9).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{}{}".format(get_column_letter(len(COLUMNS)), ws.max_row)

    validation = DataValidation(
        type="list", formula1='"compliant,partial,alternative,non-compliant"', allow_blank=True)
    ws.add_data_validation(validation)
    validation.add("G2:G{}".format(ws.max_row))

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Compliance summary"
    summary["A1"].font = Font(bold=True, size=14, color="1F3864")
    summary["A3"] = "Requirements"
    summary["B3"] = len(rows)
    summary["A4"] = "Mandatory not yet compliant"
    summary["B4"] = len(open_mandatory)
    summary["B4"].font = Font(bold=True, color="9C0006" if open_mandatory else "006100")
    summary["A5"] = "Evaluation weight unproven"
    summary["B5"] = weight_at_risk
    summary["A7"] = ("Nothing proceeds while 'mandatory not yet compliant' is above zero. "
                     "A recorded alternative counts as closed; an unnoticed gap does not.")
    summary["A7"].font = Font(italic=True, size=9, color="595959")
    summary.column_dimensions["A"].width = 46
    summary.column_dimensions["B"].width = 18

    wb.save(path)


# ---------------------------------------------------------------- cli

def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = parser.add_subparsers(dest="command")

    p_ext = sub.add_parser("extract", help="pull candidate requirements from an RFP text file")
    p_ext.add_argument("source")
    p_ext.add_argument("--out", default="requirements.csv")
    p_ext.add_argument("--min-length", type=int, default=25)
    p_ext.add_argument("--max-length", type=int, default=600)
    p_ext.set_defaults(func=run_extract)

    p_mat = sub.add_parser("matrix", help="build the compliance matrix and blocking checks")
    p_mat.add_argument("requirements")
    p_mat.add_argument("--out", help="write the markdown matrix to this path")
    p_mat.add_argument("--xlsx", help="also write an Excel compliance matrix")
    p_mat.set_defaults(func=run_matrix)

    args = parser.parse_args(argv)
    if not getattr(args, "command", None):
        parser.print_help()
        return 1
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
