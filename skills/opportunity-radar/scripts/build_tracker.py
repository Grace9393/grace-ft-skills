#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Build the bilingual opportunity tracker workbook from a JSON dataset.

    python build_tracker.py opportunities.json -o tracker.xlsx

Emits eight sheets:

    說明 Overview              bilingual key: classification, status legend, highlights
    徵件中 Open Now            only what is open or opening soon, sorted by deadline (ZH)
    Open Now (EN)              same rows, English, IDs aligned row-for-row
    社群與組織清單              partner / community directory with CRM columns (ZH)
    Communities & Orgs (EN)    same rows, English
    黑客松與競賽清單            the full opportunity directory incl. closed (ZH)
    Hackathons (EN)            same rows, English
    <legacy sheet>             optional verbatim copy of the source list, for traceability

Design rules this script enforces, because they are what make the file trustworthy:

* ZH and EN sheets are generated from ONE record each, so they cannot drift.
  Row N in the Chinese sheet is always row N in the English sheet.
* The countdown is computed against meta.verified_date, never the machine clock,
  so re-running later does not silently relabel a stale file as current.
* A record only gets a hard `deadline` when an exact date is known. Records whose
  source showed a relative countdown carry the estimate in the text and say so;
  records with no date at all sort last instead of being invented.
"""
import argparse
import io
import json
import sys
from datetime import date

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required:  pip install openpyxl")

# ---------------------------------------------------------------- palette
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")
LEGACY_FILL = PatternFill("solid", fgColor="FCE4D6")

CAT_FILL = {
    "企業": "DDEBF7", "Corporate": "DDEBF7",
    "個人": "E2EFDA", "Individual": "E2EFDA",
    "政府": "FFF2CC", "Government": "FFF2CC",
    "學研": "E4DFEC", "Academic": "E4DFEC",
    "待確認": "F2F2F2", "TBC": "F2F2F2",
}
ST_FILL = {"urgent": "FFC7CE", "open": "C6EFCE", "soon": "BDD7EE",
           "tbd": "FFF2CC", "closed": "F2F2F2", "rolling": "E4DFEC"}
ST_FONT = {"urgent": "9C0006", "open": "006100", "soon": "1F4E79",
           "tbd": "7F6000", "closed": "808080", "rolling": "5B3A87"}
ST_ZH = {"urgent": "🔴 徵件中（急）", "open": "🟢 徵件中", "soon": "🔵 即將開放",
         "tbd": "🟡 待公告", "closed": "⚪ 已截止", "rolling": "🟣 常態性"}
ST_EN = {"urgent": "🔴 Open (urgent)", "open": "🟢 Open now", "soon": "🔵 Opening soon",
         "tbd": "🟡 Not yet announced", "closed": "⚪ Closed", "rolling": "🟣 Recurring"}
ST_ORDER = {"urgent": 0, "open": 1, "soon": 2, "rolling": 3, "tbd": 4, "closed": 5}
OPEN_STATES = ("urgent", "open", "soon")

OPP_HDR_ZH = ["ID", "名稱", "狀態", "主分類", "主辦單位", "報名／徵件期限", "倒數",
              "賽事日期／流程", "形式", "地區", "主題", "參賽資格", "獎金", "連結", "來源", "備註"]
OPP_HDR_EN = ["ID", "Name", "Status", "Category", "Organizer", "Entry deadline", "Countdown",
              "Event dates / process", "Format", "Region", "Theme", "Eligibility", "Prize",
              "Link", "Source", "Notes"]
OPP_W = [6, 34, 13, 9, 30, 24, 11, 30, 20, 10, 34, 30, 26, 36, 16, 46]

ORG_HDR_ZH = ["ID", "名稱", "主分類", "細分類", "原名單類型", "範疇 (產業/範疇)", "可合作原因",
              "已有聯繫方式?", "聯繫方式", "聯繫進度", "負責人", "連結", "來源", "相關檔案"]
ORG_HDR_EN = ["ID", "Name", "Category", "Sub-category", "Original type", "Scope (industry/domain)",
              "Why partner", "Contact on file?", "Contact method", "Status", "Owner", "Link",
              "Source", "Related files"]
ORG_W = [6, 34, 10, 24, 13, 20, 50, 10, 20, 12, 15, 38, 18, 38]


def g(rec, key, default=""):
    v = rec.get(key, default)
    return default if v is None else v


def style_sheet(ws, widths, cat_col=None, st_col=None, flag_col=None, flag_values=()):
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 32
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = THIN
        if cat_col:
            v = row[cat_col - 1].value
            if v in CAT_FILL:
                row[cat_col - 1].fill = PatternFill("solid", fgColor=CAT_FILL[v])
        if st_col:
            v = row[st_col - 1].value or ""
            for key in ST_ZH:
                if v in (ST_ZH[key], ST_EN[key]):
                    row[st_col - 1].fill = PatternFill("solid", fgColor=ST_FILL[key])
                    row[st_col - 1].font = Font(bold=True, color=ST_FONT[key])
                    break
        if flag_col and row[flag_col - 1].value in flag_values:
            row[0].fill = LEGACY_FILL
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def make_countdown(verified):
    """Return (zh, en) countdown formatters bound to the verification date."""
    def days(deadline):
        if not deadline:
            return None
        try:
            y, m, d = map(int, str(deadline)[:10].split("-"))
        except ValueError:
            return None
        return (date(y, m, d) - verified).days

    def zh(deadline):
        n = days(deadline)
        if n is None:
            return "—"
        return f"剩 {n} 天" if n >= 0 else f"已過 {-n} 天"

    def en(deadline):
        n = days(deadline)
        if n is None:
            return "—"
        return f"{n} days left" if n >= 0 else f"{-n} days ago"

    return zh, en


def build(data, out_path):
    meta = data.get("meta", {})
    verified_raw = meta.get("verified_date")
    if not verified_raw:
        sys.exit("meta.verified_date is required — the countdown is meaningless without it")
    vy, vm, vd = map(int, str(verified_raw)[:10].split("-"))
    verified = date(vy, vm, vd)
    cd_zh, cd_en = make_countdown(verified)

    opps = list(data.get("opportunities", []))
    orgs = list(data.get("orgs", []))

    for o in opps:
        if o.get("st") not in ST_ZH:
            sys.exit(f"unknown status {o.get('st')!r} on {o.get('name_zh') or o.get('name_en')}; "
                     f"expected one of {sorted(ST_ZH)}")
    opps.sort(key=lambda o: (ST_ORDER[o["st"]], g(o, "deadline") or "9999-99-99"))
    for i, o in enumerate(opps, 1):
        o["_id"] = f"H{i:02d}"
    for i, o in enumerate(orgs, 1):
        o["_id"] = f"C{i:02d}"

    def opp_row(o, lang):
        z = lang == "zh"
        return [o["_id"],
                g(o, "name_zh" if z else "name_en"),
                (ST_ZH if z else ST_EN)[o["st"]],
                g(o, "cat" if z else "cat_en"),
                g(o, "host_zh" if z else "host_en"),
                g(o, "dl_zh" if z else "dl_en"),
                (cd_zh if z else cd_en)(g(o, "deadline")),
                g(o, "ev_zh" if z else "ev_en"),
                g(o, "fmt_zh" if z else "fmt_en"),
                g(o, "reg_zh" if z else "reg_en"),
                g(o, "topic_zh" if z else "topic_en"),
                g(o, "elig_zh" if z else "elig_en"),
                g(o, "prize_zh" if z else "prize_en"),
                g(o, "link"),
                g(o, "src" if z else "src_en"),
                g(o, "note_zh" if z else "note_en")]

    def org_row(c, lang):
        z = lang == "zh"
        return [c["_id"],
                g(c, "name_zh" if z else "name_en"),
                g(c, "cat" if z else "cat_en"),
                g(c, "sub_zh" if z else "sub_en"),
                g(c, "legacy_type" if z else "legacy_type_en"),
                g(c, "scope_zh" if z else "scope_en"),
                g(c, "why_zh" if z else "why_en"),
                g(c, "has_contact"),
                g(c, "contact_zh" if z else "contact_en"),
                g(c, "status"),
                g(c, "owner"),
                g(c, "link"),
                g(c, "src" if z else "src_en"),
                g(c, "files")]

    wb = Workbook()
    legacy_tags = {g(meta, "legacy_source_tag", "__none__"),
                   g(meta, "legacy_source_tag_en", "__none__")}

    # ---- Overview
    ws = wb.active
    ws.title = "說明 Overview"
    ws.append(["項目 Item", "中文說明", "English"])
    open_opps = [o for o in opps if o["st"] in OPEN_STATES]
    rows = [
        ["標題 Title", g(meta, "title_zh"), g(meta, "title_en")],
        ["版本 Version", g(meta, "version"), g(meta, "version_en") or g(meta, "version")],
        ["查證日 Verified", str(verified_raw),
         f"{verified_raw} — all countdowns are measured from this date"],
        ["資料來源 Sources", g(meta, "sources_zh"), g(meta, "sources_en")],
        ["", "", ""],
        ["分類定義 Classification", "", ""],
        ["企業 Corporate", "由公司、財團法人、協會或其生態系主導",
         "Led by a company, foundation, association or its ecosystem"],
        ["個人 Individual", "由個人創作者、志工社群、學生團隊主導",
         "Led by individual creators, volunteer communities or student teams"],
        ["政府 Government", "由中央或地方政府機關主辦、主管或出資",
         "Hosted, supervised or funded by central or local government"],
        ["學研 Academic", "由大專校院或學術單位主辦",
         "Hosted by universities or academic institutions"],
        ["", "", ""],
        ["狀態定義 Status", "", ""],
        [ST_ZH["urgent"], "30 天內截止，需立即行動", "Closing within 30 days — act now"],
        [ST_ZH["open"], "目前開放報名／徵件", "Currently accepting entries"],
        [ST_ZH["soon"], "已公告開放日期，尚未開始收件",
         "Opening date announced, entries not yet accepted"],
        [ST_ZH["tbd"], "下一屆尚未公告，需持續追蹤",
         "Next edition not yet announced — keep tracking"],
        [ST_ZH["closed"], "本屆已截止，保留供下一屆提早準備",
         "Closed — kept so the next edition can be planned early"],
        [ST_ZH["rolling"], "常態或系列舉辦，隨時可參與",
         "Recurring or series-based — joinable at any time"],
        ["", "", ""],
        ["工作表 Sheets", "", ""],
        ["徵件中 Open Now", f"{len(open_opps)} 筆，依截止日排序（中文）",
         f"{len(open_opps)} records, sorted by deadline (Chinese)"],
        ["Open Now (EN)", "同上，英文版，ID 逐筆對應", "Same records in English, IDs aligned"],
        ["社群與組織清單", f"{len(orgs)} 筆，含 CRM 欄位",
         f"{len(orgs)} records with CRM columns"],
        ["Communities & Orgs (EN)", "同上，英文版", "Same records in English"],
        ["黑客松與競賽清單", f"{len(opps)} 筆完整名錄（含已截止）",
         f"{len(opps)} records — full directory including closed events"],
        ["Hackathons (EN)", "同上，英文版", "Same records in English"],
        ["", "", ""],
        ["顏色說明 Colour key",
         "第一欄橘底＝沿用自既有名單；主分類欄 藍=企業／綠=個人／黃=政府／紫=學研",
         "Orange in column A = carried over from the pre-existing list; category column "
         "blue=Corporate, green=Individual, yellow=Government, purple=Academic"],
    ]
    for h in meta.get("highlights", []):
        rows.append([g(h, "label_zh") or g(h, "label_en"), g(h, "zh"), g(h, "en")])
    for r in rows:
        ws.append(r)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 72
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
    for r in range(2, ws.max_row + 1):
        label = ws.cell(r, 1).value or ""
        if label in ("分類定義 Classification", "狀態定義 Status", "工作表 Sheets"):
            ws.cell(r, 1).font = Font(bold=True, size=12)
        if label.startswith("★") or label.startswith("⚠"):
            for c in range(1, 4):
                ws.cell(r, c).font = Font(bold=True, color="C00000")
        if label in CAT_FILL:
            ws.cell(r, 1).fill = PatternFill("solid", fgColor=CAT_FILL[label])
        for key in ST_ZH:
            if label == ST_ZH[key]:
                ws.cell(r, 1).fill = PatternFill("solid", fgColor=ST_FILL[key])
                ws.cell(r, 1).font = Font(bold=True, color=ST_FONT[key])

    # ---- Open Now (ZH / EN)
    for title, lang, hdr in (("徵件中 Open Now", "zh", OPP_HDR_ZH),
                             ("Open Now (EN)", "en", OPP_HDR_EN)):
        s = wb.create_sheet(title)
        s.append(hdr)
        for o in open_opps:
            s.append(opp_row(o, lang))
        style_sheet(s, OPP_W, cat_col=4, st_col=3)

    # ---- Orgs (ZH / EN)
    for title, lang, hdr in (("社群與組織清單", "zh", ORG_HDR_ZH),
                             ("Communities & Orgs (EN)", "en", ORG_HDR_EN)):
        s = wb.create_sheet(title)
        s.append(hdr)
        for c in orgs:
            s.append(org_row(c, lang))
        style_sheet(s, ORG_W, cat_col=3, flag_col=13, flag_values=legacy_tags)

    # ---- Full directory (ZH / EN)
    for title, lang, hdr in (("黑客松與競賽清單", "zh", OPP_HDR_ZH),
                             ("Hackathons (EN)", "en", OPP_HDR_EN)):
        s = wb.create_sheet(title)
        s.append(hdr)
        for o in opps:
            s.append(opp_row(o, lang))
        style_sheet(s, OPP_W, cat_col=4, st_col=3)

    # ---- optional verbatim legacy sheet
    legacy = data.get("legacy_sheet")
    if legacy and legacy.get("rows"):
        s = wb.create_sheet(legacy.get("title", "原始名單")[:31])
        note = legacy.get("note")
        if note:
            s.append([note])
        s.append(legacy["header"])
        for r in legacy["rows"]:
            s.append(r)
        widths = legacy.get("widths") or [24] * len(legacy["header"])
        for c, w in enumerate(widths, 1):
            s.column_dimensions[get_column_letter(c)].width = w
        hdr_row = 2 if note else 1
        for cell in s[hdr_row]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        for row in s.iter_rows(min_row=hdr_row + 1):
            for cell in row:
                cell.alignment = WRAP
                cell.border = THIN
        if note:
            s["A1"].font = Font(bold=True, italic=True, color="C00000")

    wb.save(out_path)
    return wb, opps, orgs, open_opps


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("dataset", help="input JSON")
    ap.add_argument("-o", "--out", default="tracker.xlsx")
    args = ap.parse_args()

    with open(args.dataset, encoding="utf-8") as fh:
        data = json.load(fh)

    wb, opps, orgs, open_opps = build(data, args.out)

    print(f"wrote {args.out}")
    print(f"sheets: {wb.sheetnames}")
    print(f"opportunities: {len(opps)}  (open/soon: {len(open_opps)})   orgs: {len(orgs)}")
    counts = {}
    for o in opps:
        counts[o["st"]] = counts.get(o["st"], 0) + 1
    print("by status:", counts)
    print("\n--- open now ---")
    for o in open_opps:
        print(f"  {o['_id']}  {ST_ZH[o['st']]:12} {g(o, 'dl_zh'):32} "
              f"{g(o, 'reg_zh'):6} {g(o, 'name_zh') or g(o, 'name_en')}")


if __name__ == "__main__":
    main()
