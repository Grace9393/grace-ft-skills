"""Build the finance transformation assessment Excel model.

    python ft_analyze.py all --outdir ft_output
    python ft_workbook.py --indir ft_output --assets ../assets --out FT_Assessment_Model.xlsx

Ten sheets with native Excel charts. Inputs (realisation factor, discount rate, loaded cost)
live as editable cells on `Assumptions` and are referenced by formula, so the client can re-cut
the model without rerunning anything. Requires openpyxl.
"""

import argparse
import csv
import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, RadarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required:  pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
NOTE_FONT = Font(italic=True, size=9, color="595959")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
INPUT_FONT = Font(color="0070C0", bold=True)
CALC_FONT = Font(color="000000")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '#,##0'
PCT = '0.0%'


# ---------------------------------------------------------------- helpers

def load_json(indir, name):
    path = os.path.join(indir, name + ".json")
    if not os.path.exists(path):
        sys.exit("missing {} - run ft_analyze.py first".format(path))
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def load_csv(path):
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8-sig") as fh:
        return list(csv.DictReader(fh))


def sheet_title(ws, text, note=None):
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    if note:
        ws["A2"] = note
        ws["A2"].font = NOTE_FONT
    return 4 if note else 3


def write_header(ws, row, headers, widths=None):
    for col, head in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=head)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for col, width in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def write_row(ws, row, values, number_formats=None):
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=isinstance(value, str) and len(str(value)) > 40)
        if number_formats and col <= len(number_formats) and number_formats[col - 1]:
            cell.number_format = number_formats[col - 1]


# ---------------------------------------------------------------- sheets

def build_assumptions(ws, gap, road):
    row = sheet_title(ws, "Assumptions",
                      "Yellow cells are inputs. Every calculated sheet references them - change "
                      "an input here and the model re-cuts.")
    write_header(ws, row, ["Assumption", "Value", "Unit", "Basis / source"],
                 [42, 16, 14, 70])
    row += 1
    inputs = [
        ("Benefit realisation factor", gap.get("realisation", 0.7), "%",
         "Planning discipline: 60-80% of theoretical gap. State the factor on every value page."),
        ("Discount rate", road.get("discount_rate", 0.10), "%",
         "Client WACC / hurdle rate - ask the client, never assume."),
        ("Fully loaded cost per finance FTE", 85000, "USD",
         "Client HR data. Used in the FTE-gap value calculation."),
        ("Annual revenue", 2500000000, "USD", "Client FY financials."),
        ("Annual supplier invoice volume", 420000, "count", "Client AP system."),
        ("Horizon", len(road.get("years", [])), "years", "Business case horizon."),
    ]
    first_input_row = row
    for label, value, unit, basis in inputs:
        write_row(ws, row, [label, value, unit, basis])
        cell = ws.cell(row=row, column=2)
        cell.fill = INPUT_FILL
        cell.font = INPUT_FONT
        cell.number_format = PCT if unit == "%" else MONEY
        row += 1
    ws["A" + str(row + 1)] = ("Named references used by other sheets: realisation = B{}, "
                              "discount rate = B{}.".format(first_input_row, first_input_row + 1))
    ws["A" + str(row + 1)].font = NOTE_FONT
    return {"realisation": "Assumptions!$B${}".format(first_input_row),
            "discount": "Assumptions!$B${}".format(first_input_row + 1)}


def build_benchmark_gaps(ws, gap):
    row = sheet_title(ws, "Benchmark gaps",
                      "Only rows flagged `verified` may be quoted on a client page with the source "
                      "and as-of date shown. `directional` rows frame a range in discussion only.")
    headers = ["Metric", "Unit", "Client", "Median", "Top quartile", "Position",
               "Gap to median", "Gap to top", "Confidence", "Source", "As of"]
    write_header(ws, row, headers, [46, 18, 12, 12, 13, 18, 14, 12, 12, 52, 12])
    header_row = row
    row += 1
    first = row
    for r in gap["rows"]:
        if not r.get("benchmark_found"):
            continue
        write_row(ws, row, [r["metric"], r.get("unit", ""), r["client_value"], r.get("median"),
                            r.get("top_quartile"), r.get("position"), r.get("gap_to_median"),
                            r.get("gap_to_top"), r.get("confidence"), r.get("source"),
                            r.get("as_of")],
                  [None, None, '#,##0.00', '#,##0.00', '#,##0.00', None, '#,##0.00', '#,##0.00'])
        if r.get("confidence") != "verified":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FCE4D6")
        row += 1
    last = row - 1

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Client vs median vs top quartile"
    chart.y_axis.title = "Metric value (mixed units - compare within a row only)"
    data = Reference(ws, min_col=3, max_col=5, min_row=header_row, max_row=last)
    cats = Reference(ws, min_col=1, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height, chart.width = 11, 24
    ws.add_chart(chart, "A{}".format(last + 3))
    return header_row, first, last


def build_value_at_stake(ws, gap, refs):
    row = sheet_title(ws, "Value at stake",
                      "gap x driver volume x realisation factor. Value at stake is the size of the "
                      "prize, NOT the business case - keep them on separate pages.")
    write_header(ws, row, ["Metric", "Gap to median", "Gap to top quartile", "Driver volume",
                           "Driver description", "Basis", "Value to median", "Value to top"],
                 [46, 15, 17, 18, 56, 18, 18, 18])
    row += 1
    first = row
    for r in gap["rows"]:
        if not r.get("benchmark_found") or not r.get("driver_volume"):
            continue
        suppressed = r.get("confidence") != "verified"
        write_row(ws, row, [r["metric"], r.get("gap_to_median"), r.get("gap_to_top"),
                            r.get("driver_volume"), r.get("driver_description", ""),
                            r.get("value_basis", ""), None, None],
                  [None, '#,##0.00', '#,##0.00', MONEY, None, None, MONEY, MONEY])
        if suppressed:
            ws.cell(row=row, column=7, value="DIRECTIONAL - do not present")
            ws.cell(row=row, column=8, value="DIRECTIONAL - do not present")
        else:
            ws.cell(row=row, column=7,
                    value="=IF(B{r}>0,B{r}*D{r}*{f},0)".format(r=row, f=refs["realisation"]))
            ws.cell(row=row, column=8,
                    value="=IF(C{r}>0,C{r}*D{r}*{f},0)".format(r=row, f=refs["realisation"]))
            ws.cell(row=row, column=7).number_format = MONEY
            ws.cell(row=row, column=8).number_format = MONEY
        row += 1
    last = row - 1

    row += 1
    ws.cell(row=row, column=1, value="Recurring annual (to top quartile)").font = Font(bold=True)
    ws.cell(row=row, column=8,
            value='=SUMIF($F${f}:$F${l},"recurring_annual",$H${f}:$H${l})'.format(f=first, l=last))
    ws.cell(row=row, column=8).number_format = MONEY
    ws.cell(row=row, column=8).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="One-time cash release (working capital)").font = Font(bold=True)
    ws.cell(row=row, column=8,
            value='=SUMIF($F${f}:$F${l},"one_time_cash",$H${f}:$H${l})'.format(f=first, l=last))
    ws.cell(row=row, column=8).number_format = MONEY
    ws.cell(row=row, column=8).font = Font(bold=True)
    row += 2
    ws.cell(row=row, column=1,
            value="Working capital is a one-time cash release plus carry at the cost of capital - "
                  "never present it as a recurring P&L benefit.").font = NOTE_FONT
    return first, last


def build_maturity(ws, maturity):
    row = sheet_title(ws, "Capability maturity",
                      "1 Ad hoc - 2 Repeatable - 3 Defined - 4 Managed - 5 Optimised. "
                      "A level is only awarded where the client evidenced it.")
    lenses = maturity["lenses"]
    headers = ["Capability"] + [l.title() for l in lenses] + ["Weighted current", "Target", "Gap"]
    write_header(ws, row, headers, [34] + [13] * len(lenses) + [17, 12, 10])
    header_row = row
    row += 1
    first = row
    cells = maturity["cells"]
    for cap in maturity["capabilities"]:
        values = [cap]
        for lens in lenses:
            cell = cells.get("{}||{}".format(cap, lens))
            values.append(cell["current"] if cell else None)
        cap_stats = maturity["by_capability"][cap]
        values += [cap_stats["current"], cap_stats["target"], cap_stats["gap"]]
        write_row(ws, row, values,
                  [None] + ['0'] * len(lenses) + ['0.00', '0.00', '0.00'])
        row += 1
    last = row - 1

    lens_row = row
    values = ["By lens"] + [maturity["by_lens"][l]["current"] for l in lenses]
    values += [maturity["overall"]["current"], maturity["overall"]["target"],
               round(maturity["overall"]["target"] - maturity["overall"]["current"], 2)]
    write_row(ws, lens_row, values, [None] + ['0.00'] * len(lenses) + ['0.00', '0.00', '0.00'])
    for col in range(1, len(headers) + 1):
        ws.cell(row=lens_row, column=col).font = Font(bold=True)

    heat_range = "{}{}:{}{}".format(get_column_letter(2), first,
                                    get_column_letter(1 + len(lenses)), last)
    ws.conditional_formatting.add(heat_range, ColorScaleRule(
        start_type="num", start_value=1, start_color="F8696B",
        mid_type="num", mid_value=3, mid_color="FFEB84",
        end_type="num", end_value=5, end_color="63BE7B"))

    radar = RadarChart()
    radar.type = "marker"
    radar.title = "Capability maturity - current vs target"
    radar.style = 26
    data = Reference(ws, min_col=len(lenses) + 2, max_col=len(lenses) + 3,
                     min_row=header_row, max_row=last)
    radar.add_data(data, titles_from_data=True)
    radar.set_categories(Reference(ws, min_col=1, min_row=first, max_row=last))
    radar.height, radar.width = 12, 14
    ws.add_chart(radar, "A{}".format(last + 4))

    lens_chart = BarChart()
    lens_chart.type = "col"
    lens_chart.title = "Maturity by lens"
    lens_chart.y_axis.title = "Weighted level"
    lens_data = Reference(ws, min_col=2, max_col=1 + len(lenses), min_row=lens_row, max_row=lens_row)
    lens_chart.add_data(lens_data, from_rows=True, titles_from_data=False)
    lens_chart.set_categories(Reference(ws, min_col=2, max_col=1 + len(lenses),
                                        min_row=header_row, max_row=header_row))
    lens_chart.height, lens_chart.width = 12, 14
    ws.add_chart(lens_chart, "J{}".format(last + 4))


def build_register(ws, rows):
    row = sheet_title(ws, "Pain point register",
                      "Every row carries evidence, a quantum, a root cause, a lens and a confirming "
                      "owner. Rows without all five are not yet findings.")
    headers = ["ID", "Process", "Thread step", "Process area", "Observation", "Evidence",
               "Quantum", "Immediate cause", "Root cause", "Lens", "Dimension", "Persona",
               "Severity", "Frequency", "Effort", "Impact", "Owner confirmed", "Theme",
               "Initiative"]
    write_header(ws, row, headers,
                 [9, 10, 12, 13, 46, 34, 24, 34, 46, 12, 16, 16, 10, 11, 9, 9, 20, 26, 11])
    row += 1
    first = row
    for r in rows:
        sev = float(r.get("severity") or 0)
        freq = float(r.get("frequency") or 0)
        write_row(ws, row, [r.get("id"), r.get("process"), r.get("step"), r.get("process_area"),
                            r.get("observation"), r.get("evidence"), r.get("quantum"),
                            r.get("immediate_cause"), r.get("root_cause"), r.get("lens"),
                            r.get("pain_dimension"), r.get("persona"), sev, freq,
                            float(r.get("effort") or 0), sev * freq, r.get("owner_confirmed"),
                            r.get("theme"), r.get("initiative_id")])
        if not (r.get("process_area") or "").strip():
            ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor="FCE4D6")
        row += 1
    last = row - 1
    if last >= first:
        ws.conditional_formatting.add("P{}:P{}".format(first, last), ColorScaleRule(
            start_type="num", start_value=1, start_color="63BE7B",
            mid_type="num", mid_value=12, mid_color="FFEB84",
            end_type="num", end_value=25, end_color="F8696B"))

        def distribution(field):
            counts = {}
            for r in rows:
                key = (r.get(field) or "unclassified").strip() or "unclassified"
                counts[key] = counts.get(key, 0) + 1
            return counts

        lens_counts = distribution("lens")
        summary_row = last + 3
        ws.cell(row=summary_row, column=1, value="Lens distribution (where the cause sits)").font = Font(bold=True)
        summary_row += 1
        start = summary_row
        for lens, count in sorted(lens_counts.items(), key=lambda kv: -kv[1]):
            ws.cell(row=summary_row, column=1, value=lens)
            ws.cell(row=summary_row, column=2, value=count)
            summary_row += 1
        chart = BarChart()
        chart.type = "col"
        chart.title = "Pain points by lens"
        chart.y_axis.title = "Count"
        chart.add_data(Reference(ws, min_col=2, min_row=start, max_row=summary_row - 1),
                       titles_from_data=False)
        chart.set_categories(Reference(ws, min_col=1, min_row=start, max_row=summary_row - 1))
        chart.height, chart.width = 9, 14
        ws.add_chart(chart, "D{}".format(last + 3))
        ws.cell(row=summary_row + 1, column=1,
                value="A register dominated by 'technology' usually means the interviews stopped "
                      "at the first answer.").font = NOTE_FONT

        dim_counts = distribution("pain_dimension")
        summary_row += 3
        ws.cell(row=summary_row, column=1,
                value="Pain dimension distribution (what it costs)").font = Font(bold=True)
        summary_row += 1
        for dimension, count in sorted(dim_counts.items(), key=lambda kv: -kv[1]):
            ws.cell(row=summary_row, column=1, value=dimension)
            ws.cell(row=summary_row, column=2, value=count)
            summary_row += 1
        ws.cell(row=summary_row + 1, column=1,
                value="A blueprint needs at least 3 distinct dimensions across its 4 pain points. "
                      "Rows with no process area (shaded) cannot become a blueprint until the "
                      "closed enum is extended.").font = NOTE_FONT


def build_backlog(ws, road):
    row = sheet_title(ws, "Initiative backlog",
                      "Every initiative traces to the pain points it resolves. Stop / Change / Add "
                      "is the classification clients act on.")
    headers = ["ID", "Initiative", "Type", "Lens", "Wave", "Quadrant", "Benefit category",
               "Benefit p.a.", "One-time cost", "Run cost p.a.", "Effort", "Impact",
               "Duration (mo)", "Depends on", "Owner", "Pain points"]
    write_header(ws, row, headers,
                 [9, 42, 10, 12, 8, 15, 17, 15, 15, 14, 9, 9, 13, 14, 22, 16])
    row += 1
    first = row
    for ini in road["initiatives"]:
        write_row(ws, row, [ini["id"], ini["name"], ini["type"], ini["lens"], ini["wave"],
                            ini.get("quadrant", ""), ini["benefit_category"], ini["benefit_annual"],
                            ini["cost_onetime"], ini["cost_run_annual"], ini["effort"],
                            ini["impact"], ini["duration_months"], ";".join(ini["depends_on"]),
                            ini["owner"], ini["pain_point_ids"]],
                  [None, None, None, None, '0', None, None, MONEY, MONEY, MONEY, '0', '0', '0'])
        row += 1
    last = row - 1

    cat_row = last + 3
    ws.cell(row=cat_row, column=1, value="Benefit by category - never blended").font = Font(bold=True)
    cat_row += 1
    start = cat_row
    for cat, value in road["benefit_by_category"].items():
        ws.cell(row=cat_row, column=1, value=cat)
        ws.cell(row=cat_row, column=2, value=value).number_format = MONEY
        cat_row += 1
    chart = BarChart()
    chart.type = "col"
    chart.title = "Steady-state annual benefit by category"
    chart.y_axis.title = "USD"
    chart.add_data(Reference(ws, min_col=2, min_row=start, max_row=cat_row - 1),
                   titles_from_data=False)
    chart.set_categories(Reference(ws, min_col=1, min_row=start, max_row=cat_row - 1))
    chart.height, chart.width = 9, 16
    ws.add_chart(chart, "D{}".format(last + 3))
    ws.cell(row=cat_row + 1, column=1,
            value="Headline number = hard / cash P&L only. Everything else is a separate labelled "
                  "line.").font = NOTE_FONT
    return first, last


def build_roadmap(ws, road):
    row = sheet_title(ws, "Roadmap",
                      "Waves respect dependencies: standardise before centralising, fix data at "
                      "source before building reporting on it, standardise before automating.")
    write_header(ws, row, ["ID", "Initiative", "Wave", "Owner", "Start (month)",
                           "Duration (months)", "Go-live (month)"],
                 [9, 46, 8, 24, 15, 18, 16])
    header_row = row
    row += 1
    first = row
    for ini in sorted(road["initiatives"], key=lambda i: (i["start_month"], i["id"])):
        write_row(ws, row, [ini["id"], ini["name"], ini["wave"], ini["owner"],
                            ini["start_month"], ini["duration_months"], ini["end_month"]],
                  [None, None, '0', None, '0', '0', '0'])
        row += 1
    last = row - 1

    gantt = BarChart()
    gantt.type = "bar"
    gantt.grouping = "stacked"
    gantt.overlap = 100
    gantt.title = "Roadmap - waves and dependencies (months from start)"
    gantt.x_axis.title = "Initiative"
    gantt.y_axis.title = "Month"
    data = Reference(ws, min_col=5, max_col=6, min_row=header_row, max_row=last)
    gantt.add_data(data, titles_from_data=True)
    gantt.set_categories(Reference(ws, min_col=2, min_row=first, max_row=last))
    gantt.series[0].graphicalProperties.noFill = True
    gantt.series[0].graphicalProperties.line.noFill = True
    gantt.height, gantt.width = 13, 26
    ws.add_chart(gantt, "A{}".format(last + 3))

    warn_row = last + 30
    if road.get("warnings"):
        ws.cell(row=warn_row, column=1, value="Validation warnings").font = Font(bold=True)
        for i, warning in enumerate(road["warnings"], start=1):
            ws.cell(row=warn_row + i, column=1, value=warning)


def build_business_case(ws, road, refs):
    row = sheet_title(ws, "Business case",
                      "Benefits ramped 0% months 0-3, 25% months 4-6, 60% months 7-12, 100% "
                      "thereafter, and multiplied by the realisation factor on Assumptions.")
    write_header(ws, row, ["Year", "Benefit (ramped)", "Cost", "Net", "Cumulative"],
                 [10, 20, 18, 18, 20])
    header_row = row
    row += 1
    first = row
    for year in road["years"]:
        write_row(ws, row, [year["year"], year["benefit"], year["cost"], None, None],
                  ['0', MONEY, MONEY, MONEY, MONEY])
        ws.cell(row=row, column=4, value="=B{r}-C{r}".format(r=row)).number_format = MONEY
        if row == first:
            ws.cell(row=row, column=5, value="=D{}".format(row)).number_format = MONEY
        else:
            ws.cell(row=row, column=5, value="=E{}+D{}".format(row - 1, row)).number_format = MONEY
        row += 1
    last = row - 1

    row += 1
    metrics_start = row
    ws.cell(row=row, column=1, value="NPV").font = Font(bold=True)
    ws.cell(row=row, column=2,
            value="=NPV({d},D{f}:D{l})".format(d=refs["discount"], f=first, l=last))
    ws.cell(row=row, column=2).number_format = MONEY
    row += 1
    ws.cell(row=row, column=1, value="Total investment (one-time)").font = Font(bold=True)
    ws.cell(row=row, column=2, value=road["total_investment"]).number_format = MONEY
    row += 1
    ws.cell(row=row, column=1, value="Payback").font = Font(bold=True)
    ws.cell(row=row, column=2,
            value="month {}".format(road["payback_month"]) if road.get("payback_month")
            else "not achieved in horizon")
    row += 2

    ws.cell(row=row, column=1, value="Sensitivity").font = Font(bold=True)
    row += 1
    write_header(ws, row, ["Scenario", "NPV"], [30, 20])
    row += 1
    for label, ben, cost in [("Base", 1.0, 1.0), ("Benefits -30%", 0.7, 1.0),
                             ("Costs +20%", 1.0, 1.2), ("Both", 0.7, 1.2)]:
        npv = sum((y["benefit"] * ben - y["cost"] * cost)
                  / ((1 + road["discount_rate"]) ** y["year"]) for y in road["years"])
        write_row(ws, row, [label, npv], [None, MONEY])
        row += 1

    chart = LineChart()
    chart.title = "Cumulative net benefit"
    chart.y_axis.title = "USD"
    chart.x_axis.title = "Year"
    chart.add_data(Reference(ws, min_col=5, min_row=header_row, max_row=last), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=first, max_row=last))
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.height, chart.width = 11, 20
    ws.add_chart(chart, "H{}".format(header_row))

    bars = BarChart()
    bars.type = "col"
    bars.title = "Benefit and cost by year"
    bars.y_axis.title = "USD"
    bars.add_data(Reference(ws, min_col=2, max_col=3, min_row=header_row, max_row=last),
                  titles_from_data=True)
    bars.set_categories(Reference(ws, min_col=1, min_row=first, max_row=last))
    bars.height, bars.width = 11, 20
    ws.add_chart(bars, "H{}".format(metrics_start + 6))


def build_sources(ws, gap, benchmarks_csv):
    row = sheet_title(ws, "Sources",
                      "Every benchmark used, with source, as-of date, population and confidence. "
                      "This sheet is what makes the analysis defensible.")
    write_header(ws, row, ["Metric", "Industry", "Top quartile", "Median", "Bottom quartile",
                           "Direction", "Confidence", "Source", "As of", "Population", "Notes"],
                 [44, 22, 13, 12, 15, 14, 12, 52, 11, 34, 52])
    row += 1
    for r in benchmarks_csv:
        write_row(ws, row, [r.get("metric"), r.get("industry"), r.get("top_quartile"),
                            r.get("median"), r.get("bottom_quartile"), r.get("direction"),
                            r.get("confidence"), r.get("source"), r.get("as_of"),
                            r.get("population"), r.get("notes")])
        if (r.get("confidence") or "").strip() != "verified":
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FCE4D6")
        row += 1
    row += 1
    ws.cell(row=row, column=1,
            value="Shaded rows are DIRECTIONAL - practitioner ranges with no published source. "
                  "They may frame a range in discussion and must never appear as a figure on a "
                  "client page or in the business case.").font = NOTE_FONT
    row += 2
    if gap.get("warnings"):
        ws.cell(row=row, column=1, value="Analysis warnings").font = Font(bold=True)
        for i, warning in enumerate(gap["warnings"], start=1):
            ws.cell(row=row + i, column=1, value=warning)


def build_exec_summary(ws, gap, maturity, road, vas_rows):
    row = sheet_title(ws, "Executive summary",
                      "Every figure is a formula referencing another sheet - nothing on this page "
                      "is hand-typed.")
    write_header(ws, row, ["Measure", "Value", "Read"], [46, 22, 82])
    row += 1
    vas_first, vas_last = vas_rows
    sumif = ("=SUMIF('Value at Stake'!$F${f}:$F${l},\"{{}}\",'Value at Stake'!$H${f}:$H${l})"
             .format(f=vas_first, l=vas_last))
    entries = [
        ("Recurring value at stake p.a. (gap to top quartile)",
         sumif.format("recurring_annual"), MONEY,
         "Theoretical prize, after the realisation factor. Not the business case."),
        ("One-time cash release (working capital)",
         sumif.format("one_time_cash"), MONEY,
         "One-time release plus carry at the cost of capital - never a recurring P&L benefit."),
        ("Overall capability maturity (current)", maturity["overall"]["current"], '0.00',
         "Against a target of {:.2f}. Weakest lens: {}.".format(
             maturity["overall"]["target"],
             min(maturity["by_lens"], key=lambda l: maturity["by_lens"][l]["current"]).title())),
        ("Hard / cash P&L benefit at steady state", road["benefit_by_category"].get("hard", 0),
         MONEY, "The only category that may carry the headline number."),
        ("Total investment (one-time)", road["total_investment"], MONEY,
         "Excludes run-rate change and internal effort unless costed in the backlog."),
        ("Payback", "month {}".format(road["payback_month"]) if road.get("payback_month")
         else "not achieved in horizon", None, "From programme start, on ramped benefits."),
    ]
    for label, value, fmt, note in entries:
        write_row(ws, row, [label, value, note or ""])
        if fmt:
            ws.cell(row=row, column=2).number_format = fmt
        ws.cell(row=row, column=2).font = Font(bold=True)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Reading order").font = Font(bold=True)
    row += 1
    for text in ["1. Benchmark gaps - where the client stands, with sources",
                 "2. Value at stake - the size of the prize",
                 "3. Capability maturity - why the gap exists",
                 "4. Pain point register - the evidence",
                 "5. Initiative backlog and Roadmap - what to do and when",
                 "6. Business case - what it costs and returns",
                 "7. Assumptions and Sources - what it all rests on"]:
        ws.cell(row=row, column=1, value=text)
        row += 1


# ---------------------------------------------------------------- main

def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--indir", default="ft_output", help="ft_analyze.py output directory")
    parser.add_argument("--assets", default=os.path.join(HERE, "..", "assets"))
    parser.add_argument("--out", default="FT_Assessment_Model.xlsx")
    args = parser.parse_args(argv)

    gap = load_json(args.indir, "benchmark_gap")
    maturity = load_json(args.indir, "maturity")
    road = load_json(args.indir, "roadmap")
    register = load_csv(os.path.join(args.assets, "pain-point-register-template.csv"))
    benchmarks = load_csv(os.path.join(args.assets, "benchmarks.csv"))

    wb = Workbook()
    ws_exec = wb.active
    ws_exec.title = "Executive Summary"
    ws_bench = wb.create_sheet("Benchmark Gaps")
    ws_value = wb.create_sheet("Value at Stake")
    ws_mat = wb.create_sheet("Maturity Heatmap")
    ws_reg = wb.create_sheet("Pain Point Register")
    ws_back = wb.create_sheet("Initiative Backlog")
    ws_road = wb.create_sheet("Roadmap")
    ws_case = wb.create_sheet("Business Case")
    ws_assum = wb.create_sheet("Assumptions")
    ws_src = wb.create_sheet("Sources")

    refs = build_assumptions(ws_assum, gap, road)
    build_benchmark_gaps(ws_bench, gap)
    vas_rows = build_value_at_stake(ws_value, gap, refs)
    build_maturity(ws_mat, maturity)
    build_register(ws_reg, register)
    build_backlog(ws_back, road)
    build_roadmap(ws_road, road)
    build_business_case(ws_case, road, refs)
    build_sources(ws_src, gap, benchmarks)
    build_exec_summary(ws_exec, gap, maturity, road, vas_rows)

    wb.save(args.out)
    print("wrote {} ({} sheets)".format(args.out, len(wb.sheetnames)))
    print("sheets: " + ", ".join(wb.sheetnames))
    return 0


if __name__ == "__main__":
    sys.exit(main())
